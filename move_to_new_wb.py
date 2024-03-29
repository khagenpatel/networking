import csv
from openpyxl import Workbook
from datatime import datetime

timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")

# Initialize Workbook
wb = Workbook()

# The active worksheet will contain all the data
ws_all = wb.active
ws_all.title = 'All Data'

# We will move duplicates to this worksheet
ws_duplicates = wb.create_sheet(title='Duplicates')

# Open CSV file and write to Excel
with open('input.csv', mode='r') as csv_file:
    csv_reader = csv.reader(csv_file)
    headers = next(csv_reader)
    ws_all.append(headers)
    ws_duplicates.append(headers)
    data_rows = [list(row) for row in csv_reader]
    for row in data_rows:
        ws_all.append(row)

# Check for duplicates and move to "Duplicates" sheet
live_hosts = {}
rows_to_delete = []

for row in ws_all.iter_rows(min_row=2, min_col=4, max_col=4):  # Live_Host is the 4th column
    for cell in row:
        if cell.value in live_hosts:
            # If this value has been seen before, it is a duplicate and should be moved
            duplicate_row = ws_all[cell.row]
            ws_duplicates.append([c.value for c in duplicate_row])
            rows_to_delete.append(cell.row)
        else:
            # First time seeing this value, just remember it
            live_hosts[cell.value] = cell

# Delete rows from 'All Data' after moving to 'Duplicates'
for row_id in sorted(rows_to_delete, reverse=True):
    ws_all.delete_rows(row_id)

# Save the workbook to a file
wb.save(f"output-{timestamp}.xlsx")
