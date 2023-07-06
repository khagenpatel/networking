import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Load data from the merged CSV file into a pandas DataFrame
data = pd.read_csv('merged_merged_data.csv')

# Write the DataFrame to an Excel file
excel_file = 'colored_data.xlsx'
data.to_excel(excel_file, index=False)

# Load the workbook
workbook = load_workbook(excel_file)
sheet = workbook.active

# Define fills
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

# Keep track of seen mac_addresses
seen_mac_addresses = set()

# Iterate through rows and color them
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, max_col=sheet.max_column):
    mac_address_cell = row[1]  # Assuming mac_address is the second column
    mac_address = mac_address_cell.value

    if mac_address:
        if mac_address not in seen_mac_addresses:
            # If mac_address is seen for the first time, color row green
            for cell in row:
                cell.fill = green_fill
            seen_mac_addresses.add(mac_address)
        else:
            # If mac_address has been seen before, color row gray
            for cell in row:
                cell.fill = gray_fill

# Save the modified workbook
workbook.save(excel_file)
