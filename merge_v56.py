import csv
import re
from openpyxl import load_workbook

try:
    # Load the Excel file
    workbook = load_workbook('input1.xlsx')
    sheet_name = 'Your_Sheet_Name'  # Replace 'Your_Sheet_Name' with the actual name of the worksheet
    sheet = workbook[sheet_name]

    # Load the CSV file
    with open('input2.csv', 'r', encoding='utf-8-sig') as csvfile:  # Specify the encoding for the CSV file
        csvreader = csv.reader(csvfile)
        next(csvreader)  # Skip the header row

        # Create a dictionary to store the mapping between hostnames and the other columns
        csv_data = {}
        for row in csvreader:
            if len(row) > 0:
                live_host = row[0].lower().encode('ascii', 'ignore').decode()  # Convert Live_Host to lowercase and ASCII
                csv_data[live_host] = row[1:]  # Store the remaining columns in the dictionary

    # Iterate through the rows in the Excel file
    for row in sheet.iter_rows(min_row=2):
        if len(row) > 3:
            live_host_full = row[3].value  # Extract the full Live_Host value
            match = re.search(r'\b(\w+)\b', live_host_full)  # Extract the hostname using regex
            if match:
                live_host = match.group(1).lower().encode('ascii', 'ignore').decode()  # Convert Live_Host to lowercase and ASCII
                if live_host in csv_data:
                    # Get the values from the CSV file
                    values = csv_data[live_host]

                    # Add the values to the Excel file
                    try:
                        row[7].value = values[0] if len(values) > 0 else None  # Owning Transaction Cycle
                        row[8].value = values[1] if len(values) > 1 else None  # CDR
                        row[9].value = values[2] if len(values) > 2 else None  # IT Business Service
                        row[10].value = values[3] if len(values) > 3 else None  # IT Service Instance
                        row[11].value = values[4] if len(values) > 4 else None  # ITSI Environment
                    except IndexError:
                        print("IndexError: Row data:", row)
                        print("IndexError: Values:", values)
                        raise

    # Save the modified Excel file
    workbook.save('output.xlsx')

    print("Data successfully updated in the Excel file!")

except Exception as e:
    print("An error occurred:", str(e))
