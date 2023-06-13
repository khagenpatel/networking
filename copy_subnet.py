import csv
from openpyxl import Workbook
from openpyxl.styles import PatternFill

def convert_csv_to_excel(csv_file_path, excel_file_path):
    # Read the CSV file
    rows = []
    with open(csv_file_path, 'r') as file:
        reader = csv.reader(file)
        rows = list(reader)

    # Create a new workbook
    workbook = Workbook()
    sheet = workbook.active

    # Write rows to the Excel file
    for row in rows:
        sheet.append(row)

    # Highlight duplicate values
    duplicate_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    data = sheet.values
    columns = next(data)
    seen = set()
    duplicate_indices = []

    for index, row in enumerate(data, start=2):
        key = tuple(row)
        if key in seen:
            duplicate_indices.append(index)
        else:
            seen.add(key)

    for col in sheet.iter_cols(min_row=2, max_row=sheet.max_row):
        for cell in col:
            if cell.row in duplicate_indices:
                cell.fill = duplicate_fill

    # Save the Excel file
    workbook.save(excel_file_path)

    print(f"CSV file '{csv_file_path}' has been converted to Excel file '{excel_file_path}' with highlighted duplicates.")

# Specify the file paths
csv_file_path = 'your_csv_file_path.csv'
excel_file_path = 'your_excel_file_path.xlsx'

# Call the function
convert_csv_to_excel(csv_file_path, excel_file_path)
