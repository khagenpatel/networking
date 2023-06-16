import csv
from openpyxl import load_workbook

# Read the CSV into a dictionary for easy lookup
csv_dict = {}
with open('file.csv', 'r') as f:
    reader = csv.reader(f)
    next(reader)  # skip header
    for row in reader:
        csv_dict[row[0]] = {'MAC_Address': row[1], 'Company': row[2]}

# Load the Excel workbook
wb = load_workbook('file.xlsx')

# Select the first sheet in the workbook
ws = wb[wb.sheetnames[0]]

# Add new headers for MAC Address and Company
ws.cell(row=1, column=ws.max_column + 1, value="MAC_Address")
ws.cell(row=1, column=ws.max_column + 1, value="Company")

# Iterate through each row in the worksheet
for row in ws.iter_rows(min_row=2, values_only=False):  # We need the cells not the values
    live_ip = row[4].value  # Assuming 'Live_IP' is the fifth column

    # Lookup the MAC Address and Company in the dictionary
    if live_ip in csv_dict:
        mac_address = csv_dict[live_ip]['MAC_Address']
        company = csv_dict[live_ip]['Company']
    else:
        mac_address = 'not found'
        company = 'not found'

    # Write MAC Address and Company to the row
    ws.cell(row=row[0].row, column=ws.max_column - 1, value=mac_address)
    ws.cell(row=row[0].row, column=ws.max_column, value=company)

# Save workbook
wb.save('output.xlsx')
